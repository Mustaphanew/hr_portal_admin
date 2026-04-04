import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'بوابة الإدارة'
ws.sheet_view.rightToLeft = True

header_fill = PatternFill('solid', fgColor='1A3A5C')
header_font = Font(name='Cairo', bold=True, color='FFFFFF', size=11)
main_fill = PatternFill('solid', fgColor='0D9488')
main_font = Font(name='Cairo', bold=True, color='FFFFFF', size=11)
sub_white = PatternFill('solid', fgColor='FFFFFF')
sub_gray = PatternFill('solid', fgColor='F5F5F5')
sub_font = Font(name='Cairo', size=10)
done_fill = PatternFill('solid', fgColor='C8E6C9')
done_font = Font(name='Cairo', bold=True, color='2E7D32', size=10)
progress_fill = PatternFill('solid', fgColor='FFF9C4')
progress_font = Font(name='Cairo', bold=True, color='F57F17', size=10)
notstart_fill = PatternFill('solid', fgColor='FFCDD2')
notstart_font = Font(name='Cairo', bold=True, color='C62828', size=10)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='E0E0E0'),
    right=Side(style='thin', color='E0E0E0'),
    top=Side(style='thin', color='E0E0E0'),
    bottom=Side(style='thin', color='E0E0E0'))

ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 50
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 14
ws.column_dimensions['G'].width = 55

headers = ['الرقم', 'المهمة', 'النوع', 'المسؤول', 'الأولوية', 'الحالة', 'الوصف']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border

dev = 'مصطفى الأهدل'
tasks = [
    ('1', 'إعداد بيئة التطوير والمشروع', 'رئيسية', dev, 'عالية', 'مكتمل', 'تهيئة بيئة التطوير وإنشاء هيكل المشروع'),
    ('1.1', 'إنشاء مشروع Flutter وتهيئة الهيكل', 'فرعية', dev, 'عالية', 'مكتمل', 'flutter create مع Clean Architecture'),
    ('1.2', 'إعداد الحزم والتبعيات (pubspec.yaml)', 'فرعية', dev, 'عالية', 'مكتمل', 'Dio, Riverpod, GoRouter, GetIt, SecureStorage'),
    ('1.3', 'إعداد نظام التوجيه (GoRouter + ShellRoute)', 'فرعية', dev, 'عالية', 'مكتمل', 'StatefulShellRoute + bottom nav persistence'),
    ('1.4', 'إعداد حقن التبعيات (GetIt)', 'فرعية', dev, 'عالية', 'مكتمل', 'تسجيل ApiClient و Repositories'),
    ('1.5', 'إعداد App Flavors (dev/staging/prod)', 'فرعية', dev, 'عالية', 'مكتمل', 'AppConfig.fromEnvironment() + dart-define'),
    ('2', 'طبقة الشبكة والمصادقة', 'رئيسية', dev, 'عالية', 'مكتمل', 'إعداد الاتصال بالسيرفر ونظام تسجيل الدخول'),
    ('2.1', 'إعداد ApiClient مع Dio و Interceptors', 'فرعية', dev, 'عالية', 'مكتمل', 'AuthInterceptor + Debug Logging + BaseResponse'),
    ('2.2', 'تطوير شاشة تسجيل الدخول', 'فرعية', dev, 'عالية', 'مكتمل', 'username/password مع validation'),
    ('2.3', 'إدارة الجلسة (SessionManager + SecureStorage)', 'فرعية', dev, 'عالية', 'مكتمل', 'حفظ واسترجاع التوكن والبيانات'),
    ('2.4', 'شاشة البداية (Splash Screen)', 'فرعية', dev, 'متوسطة', 'مكتمل', 'التحقق من الجلسة والتوجيه التلقائي'),
    ('2.5', 'تسجيل الخروج وتنظيف الجلسة', 'فرعية', dev, 'عالية', 'مكتمل', 'invalidate جميع Providers + مسح التخزين'),
    ('2.6', 'معالجة Timeout بحذف التخزين والانتقال للدخول', 'فرعية', dev, 'عالية', 'مكتمل', 'عند TimeoutException يتم clearAll + login'),
    ('3', 'لوحة التحكم الرئيسية (Dashboard)', 'رئيسية', dev, 'عالية', 'مكتمل', 'لوحة إدارية شاملة مع مؤشرات وإجراءات سريعة'),
    ('3.1', 'SliverAppBar مع بيانات المستخدم والوقت', 'فرعية', dev, 'عالية', 'مكتمل', 'ترويسة ثابتة + وقت AM/PM + تاريخ عربي/إنجليزي'),
    ('3.2', 'المؤشرات التشغيلية (KPI Grid)', 'فرعية', dev, 'عالية', 'مكتمل', 'موظفين، حضور، غياب، تأخر، طلبات، إجازات'),
    ('3.3', 'الإجراءات السريعة (Quick Actions)', 'فرعية', dev, 'عالية', 'مكتمل', '12 إجراء سريع للوصول المباشر'),
    ('3.4', 'ملخص الأقسام مع نسب الأداء', 'فرعية', dev, 'متوسطة', 'مكتمل', 'شريط أفقي قابل للتمرير'),
    ('3.5', 'اختيار الشركة/الفرع (BranchSelector)', 'فرعية', dev, 'عالية', 'مكتمل', 'BottomSheet مع تجميع حسب الشركة + حفظ للجلسة'),
    ('3.6', 'التنبيهات (مهام متأخرة، طلبات معلقة)', 'فرعية', dev, 'متوسطة', 'مكتمل', 'AlertBanner ديناميكي'),
    ('4', 'إدارة الموظفين', 'رئيسية', dev, 'عالية', 'مكتمل', 'عرض وبحث وفلترة الموظفين'),
    ('4.1', 'شاشة قائمة الموظفين مع بحث وفلترة', 'فرعية', dev, 'عالية', 'مكتمل', 'بحث + فلتر حسب القسم/الحالة + pagination'),
    ('4.2', 'شاشة تفاصيل الموظف', 'فرعية', dev, 'عالية', 'مكتمل', 'بيانات شخصية ووظيفية كاملة'),
    ('4.3', 'فلترة حسب الشركة/الفرع المختار', 'فرعية', dev, 'عالية', 'مكتمل', 'company_id / branch_id مع selectedBranchProvider'),
    ('5', 'إدارة الأقسام', 'رئيسية', dev, 'متوسطة', 'مكتمل', 'عرض الأقسام وتفاصيلها'),
    ('5.1', 'شاشة قائمة الأقسام', 'فرعية', dev, 'متوسطة', 'مكتمل', 'عرض جميع الأقسام مع العدد'),
    ('5.2', 'شاشة تفاصيل القسم وأعضائه', 'فرعية', dev, 'متوسطة', 'مكتمل', 'قائمة موظفي القسم'),
    ('6', 'إدارة الطلبات', 'رئيسية', dev, 'عالية', 'مكتمل', 'إدارة طلبات الموظفين مع 7 حالات'),
    ('6.1', 'شاشة الطلبات مع فلاتر الحالات التفاعلية', 'فرعية', dev, 'عالية', 'مكتمل', 'معلق/الكل/قيد المعالجة/معتمد/مرفوض/مكتمل/ملغي'),
    ('6.2', 'شاشة تفاصيل الطلب مع الموافقة/الرفض', 'فرعية', dev, 'عالية', 'مكتمل', 'POST /decide مع ملاحظات'),
    ('6.3', 'شاشة جميع الطلبات (AllRequestsScreen)', 'فرعية', dev, 'عالية', 'مكتمل', 'قائمة مع FilterBar + pagination'),
    ('6.4', 'دعم 10 أنواع طلبات', 'فرعية', dev, 'عالية', 'مكتمل', 'إجازة، حضور، إذن، مصروف، سلف، وثيقة، قرض...'),
    ('7', 'إدارة المهام', 'رئيسية', dev, 'عالية', 'مكتمل', 'إدارة المهام مع إحصائيات'),
    ('7.1', 'لوحة المهام مع الإحصائيات', 'فرعية', dev, 'عالية', 'مكتمل', 'معلقة/جارية/متأخرة + أقسام'),
    ('7.2', 'شاشة جميع المهام مع فلترة', 'فرعية', dev, 'عالية', 'مكتمل', 'FilterBar + pagination'),
    ('7.3', 'شاشة تفاصيل المهمة', 'فرعية', dev, 'عالية', 'مكتمل', 'Timeline + إعادة تعيين + إكمال'),
    ('7.4', 'إنشاء/تعديل/حذف مهمة (CRUD)', 'فرعية', dev, 'عالية', 'لم يبدأ', 'واجهة إنشاء مهمة جديدة'),
    ('8', 'إدارة الحضور والانصراف', 'رئيسية', dev, 'عالية', 'مكتمل', 'متابعة حضور الموظفين يومياً'),
    ('8.1', 'شاشة الحضور اليومي مع فلاتر', 'فرعية', dev, 'عالية', 'مكتمل', 'حاضر/متأخر/غائب/إجازة + SliverAppBar'),
    ('8.2', 'شاشة تفاصيل حضور الموظف', 'فرعية', dev, 'عالية', 'مكتمل', 'سجل الدخول/الخروج + التأخير/الإضافي'),
    ('9', 'إدارة الإجازات', 'رئيسية', dev, 'عالية', 'مكتمل', 'إدارة إجازات الموظفين مع 5 حالات'),
    ('9.1', 'شاشة الإجازات مع فلاتر تفاعلية', 'فرعية', dev, 'عالية', 'مكتمل', 'معلق/الكل/مسودة/معتمد/مرفوض/ملغي + تمرير أفقي'),
    ('9.2', 'شاشة تفاصيل الإجازة مع الموافقة/الرفض', 'فرعية', dev, 'عالية', 'مكتمل', 'بيانات الموظف + مسار الموافقات + أزرار القرار'),
    ('9.3', 'الرفض يتطلب ملاحظات إجبارية', 'فرعية', dev, 'عالية', 'مكتمل', 'validation عند الرفض بدون سبب'),
    ('10', 'المتابعة', 'رئيسية', dev, 'عالية', 'مكتمل', 'متابعة المهام والطلبات'),
    ('10.1', 'شاشة المتابعة مع إحصائيات', 'فرعية', dev, 'عالية', 'مكتمل', 'إجمالي/معلق/متأخر/مصعد'),
    ('10.2', 'شاشة تفاصيل المتابعة مع التصعيد', 'فرعية', dev, 'عالية', 'مكتمل', 'تحديث الحالة + تصعيد مع السبب'),
    ('11', 'الإعلانات', 'رئيسية', dev, 'متوسطة', 'مكتمل', 'إدارة إعلانات الشركة'),
    ('11.1', 'شاشة الإعلانات مع تصنيفات', 'فرعية', dev, 'متوسطة', 'مكتمل', 'الكل/منشور/مسودة/مثبت'),
    ('11.2', 'شاشة تفاصيل الإعلان', 'فرعية', dev, 'متوسطة', 'مكتمل', 'عرض + نشر/إلغاء نشر'),
    ('12', 'إدارة المشاريع', 'رئيسية', dev, 'متوسطة', 'مكتمل', 'متابعة المشاريع والمراحل'),
    ('12.1', 'شاشة نظرة عامة على المشاريع', 'فرعية', dev, 'متوسطة', 'مكتمل', 'إحصائيات + قائمة مختصرة'),
    ('12.2', 'شاشة قائمة المشاريع مع فلترة', 'فرعية', dev, 'متوسطة', 'مكتمل', 'فلتر الحالة + pagination'),
    ('12.3', 'شاشة تفاصيل المشروع', 'فرعية', dev, 'متوسطة', 'مكتمل', 'التقدم + الفريق + الجدول الزمني'),
    ('12.4', 'مهام ومراحل المشروع', 'فرعية', dev, 'متوسطة', 'مكتمل', 'قوائم المهام + المراحل الزمنية'),
    ('12.5', 'تحليلات المشروع', 'فرعية', dev, 'منخفضة', 'مكتمل', 'نسب الإنجاز والرسوم البيانية'),
    ('13', 'إدارة المصروفات', 'رئيسية', dev, 'عالية', 'مكتمل', 'مراجعة واعتماد طلبات المصروفات'),
    ('13.1', 'شاشة نظرة عامة على المصروفات', 'فرعية', dev, 'عالية', 'مكتمل', 'إجمالي + معتمد + قيد المراجعة'),
    ('13.2', 'شاشة طلبات المصروفات مع فلترة', 'فرعية', dev, 'عالية', 'مكتمل', 'بحث + فلتر + pagination'),
    ('13.3', 'شاشة تفاصيل المصروف مع الموافقة/الرفض', 'فرعية', dev, 'عالية', 'مكتمل', 'بطاقة المبلغ + المرفقات + القرار'),
    ('13.4', 'توزيع المصروفات حسب الفئة', 'فرعية', dev, 'متوسطة', 'مكتمل', 'رسم بياني شريطي + نسب'),
    ('13.5', 'تحليلات المصروفات', 'فرعية', dev, 'منخفضة', 'مكتمل', 'اتجاهات + أعلى الأقسام إنفاقا'),
    ('14', 'التقارير والتحليلات', 'رئيسية', dev, 'متوسطة', 'مكتمل', 'تقارير الأداء والإحصائيات'),
    ('14.1', 'مؤشرات الأداء الرئيسية (KPIs)', 'فرعية', dev, 'متوسطة', 'مكتمل', 'لوحة مؤشرات شاملة'),
    ('14.2', 'اتجاه الحضور الشهري', 'فرعية', dev, 'متوسطة', 'مكتمل', 'رسم بياني 6 أشهر'),
    ('14.3', 'تحليل الإجازات حسب النوع/القسم', 'فرعية', dev, 'متوسطة', 'مكتمل', 'توزيع الإجازات'),
    ('14.4', 'إنجاز المهام حسب القسم', 'فرعية', dev, 'متوسطة', 'مكتمل', 'نسب الإنجاز لكل قسم'),
    ('15', 'الوثائق والإشعارات', 'رئيسية', dev, 'متوسطة', 'مكتمل', 'إدارة الوثائق ومركز الإشعارات'),
    ('15.1', 'شاشة الوثائق مع تصنيفات', 'فرعية', dev, 'متوسطة', 'مكتمل', 'فلترة حسب الفئة + pagination'),
    ('15.2', 'مركز الإشعارات', 'فرعية', dev, 'عالية', 'مكتمل', 'عرض + تعليم كمقروءة + فلترة'),
    ('15.3', 'Firebase Cloud Messaging (FCM)', 'فرعية', dev, 'عالية', 'مكتمل', 'استقبال إشعارات الخادم'),
    ('16', 'التصميم وتجربة المستخدم (UI/UX)', 'رئيسية', dev, 'عالية', 'مكتمل', 'واجهة احترافية مع دعم RTL والوضع الداكن'),
    ('16.1', 'نظام الألوان الديناميكي (AppColorsExtension)', 'فرعية', dev, 'عالية', 'مكتمل', 'context.appColors لدعم الوضع الفاتح/الداكن'),
    ('16.2', 'دعم الوضع الداكن في جميع الشاشات', 'فرعية', dev, 'عالية', 'مكتمل', 'ترحيل 15+ ملف من AppColors الثابتة'),
    ('16.3', 'دعم اللغة العربية والإنجليزية (i18n)', 'فرعية', dev, 'عالية', 'مكتمل', 'ar.json + en.json + .tr(context)'),
    ('16.4', 'خط Cairo المحلي (8 أوزان)', 'فرعية', dev, 'متوسطة', 'مكتمل', 'بدون Google Fonts - محلي فقط'),
    ('16.5', 'ويدجتس مشتركة (AdminWidgets)', 'فرعية', dev, 'عالية', 'مكتمل', 'أزرار، بادجات، كروت، AppBar مخصص'),
    ('17', 'الإعدادات والملف الشخصي', 'رئيسية', dev, 'متوسطة', 'مكتمل', 'إعدادات التطبيق وبيانات المدير'),
    ('17.1', 'شاشة الإعدادات (المظهر، اللغة)', 'فرعية', dev, 'متوسطة', 'مكتمل', 'تبديل داكن/فاتح + عربي/إنجليزي'),
    ('17.2', 'الملف الشخصي مع SliverAppBar', 'فرعية', dev, 'عالية', 'مكتمل', 'بيانات من GET /profile + زر تحديث'),
    ('17.3', 'تغيير كلمة المرور (ExpansionTile)', 'فرعية', dev, 'متوسطة', 'مكتمل', 'POST /change-password + تسجيل خروج'),
    ('17.4', 'الصلاحيات والأدوار (ExpansionTile)', 'فرعية', dev, 'متوسطة', 'مكتمل', 'عرض أدوار المدير'),
    ('18', 'نظام التقسيم التلقائي (Infinite Scroll)', 'رئيسية', dev, 'عالية', 'مكتمل', 'جلب 50 عنصر + تحميل المزيد عند التمرير'),
    ('18.1', 'PaginatedNotifier (قاعدة عامة)', 'فرعية', dev, 'عالية', 'مكتمل', 'AsyncNotifier مع fetchMore + guards'),
    ('18.2', 'PaginatedListView (ويدجت التمرير)', 'فرعية', dev, 'عالية', 'مكتمل', 'ScrollController + loading footer + error retry'),
    ('18.3', 'ترحيل 9 شاشات للنظام الجديد', 'فرعية', dev, 'عالية', 'مكتمل', 'طلبات/مهام/موظفين/مصروفات/إجازات/إعلانات/مشاريع/وثائق/متابعة'),
    ('19', 'النشر والتوزيع (CI/CD)', 'رئيسية', dev, 'عالية', 'قيد التنفيذ', 'نشر التطبيق على الويب والمتاجر'),
    ('19.1', 'GitHub Actions - GitHub Pages (Web)', 'فرعية', dev, 'عالية', 'مكتمل', 'deploy.yml مع staging flavor تلقائي'),
    ('19.2', 'بناء APK/AAB للأندرويد', 'فرعية', dev, 'عالية', 'لم يبدأ', 'flutter build appbundle --release'),
    ('19.3', 'بناء IPA لـ iOS', 'فرعية', dev, 'عالية', 'لم يبدأ', 'flutter build ipa --release'),
    ('19.4', 'رفع على Google Play Store', 'فرعية', dev, 'عالية', 'لم يبدأ', 'بطاقة بيانات + لقطات + سياسة خصوصية'),
    ('19.5', 'رفع على Apple App Store', 'فرعية', dev, 'عالية', 'لم يبدأ', 'App Store Connect + مراجعة'),
]

row = 2
for i, t in enumerate(tasks):
    num, name, typ, responsible, priority, status, desc = t
    is_main = typ == 'رئيسية'
    for col, val in enumerate([num, name, typ, responsible, priority, status, desc], 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.border = thin_border
        if is_main:
            cell.fill = main_fill
            cell.font = main_font
            cell.alignment = center_align if col not in (2, 7) else right_align
        else:
            cell.fill = sub_gray if (row % 2 == 0) else sub_white
            cell.font = sub_font
            cell.alignment = center_align if col in (1, 3, 4, 5) else right_align
        if col == 6:
            cell.alignment = center_align
            if not is_main:
                if status == 'مكتمل':
                    cell.fill = done_fill
                    cell.font = done_font
                elif 'قيد' in status:
                    cell.fill = progress_fill
                    cell.font = progress_font
                elif status == 'لم يبدأ':
                    cell.fill = notstart_fill
                    cell.font = notstart_font
    row += 1

row += 2
ws.cell(row=row, column=1, value='ملخص المشروع').font = Font(name='Cairo', bold=True, size=12)
row += 1
for label, val in [
    ('اسم المشروع', 'بوابة الإدارة - HR Admin Portal'),
    ('المسؤول', 'مصطفى الأهدل'),
    ('التقنيات', 'Flutter + Dart + Dio + Riverpod + GoRouter + GetIt'),
    ('الخادم', 'Laravel API (REST)'),
    ('المنصات', 'Android + iOS + Web'),
    ('اللغات', 'العربية + الإنجليزية (RTL/LTR)'),
]:
    ws.cell(row=row, column=2, value=label).font = Font(name='Cairo', bold=True, size=10)
    ws.cell(row=row, column=3, value=val).font = Font(name='Cairo', size=10)
    row += 1

output = r'C:\Users\mustapha\Documents\app_project\hr_portal_admin\hr_portal_admin\hr_admin_tasks.xlsx'
wb.save(output)
main_count = sum(1 for t in tasks if t[2] == 'رئيسية')
sub_count = sum(1 for t in tasks if t[2] == 'فرعية')
print(f'Saved: {output}')
print(f'Total: {len(tasks)} | Main: {main_count} | Sub: {sub_count}')
